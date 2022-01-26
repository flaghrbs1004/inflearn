import requests
from bs4 import BeautifulSoup
import openpyxl

wb = openpyxl.Workbook()
ws = wb.create_sheet('맴버')
ws['A1'] = '번호'
ws['B1'] = '성명'
ws['A2'] = '1'
ws['B2'] = '임호균'
ws['A3'] = '2'
ws['B3'] = '박미리'
# ws.title = 'testTitle'
wb.save(r'/Users/imhogyun/Downloads/testExcel.xlsx')

