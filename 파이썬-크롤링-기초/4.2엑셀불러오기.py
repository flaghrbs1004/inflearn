import requests
from bs4 import BeautifulSoup
import openpyxl

fpath = r'/Users/imhogyun/Downloads/testExcel.xlsx'

wb = openpyxl.load_workbook(fpath)

ws = wb['맴버']

ws['A4'] = '3'
ws['B4'] = '2세~'

wb.save(fpath)