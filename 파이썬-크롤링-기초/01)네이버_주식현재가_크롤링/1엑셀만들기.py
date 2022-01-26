import openpyxl
import getpass
from pykrx import stock

wb = openpyxl.Workbook()
ws = wb.create_sheet('현재가')

attrs = [
    '종목'
    , '현재가'
    , '평균매입가'
    , '잔고수량'
    , '평가금액'
    , '수익율'
]

for idx, val in enumerate(attrs):
    ws[f'{chr(65+idx)}1'] = val

codes = [
    '005930'
    , '000660'
    , '035720'
]
for idx, val in enumerate(codes,2):
    종목 = stock.get_market_ticker_name(val)
    ws[f'A{idx}'] = 종목

wb.save('testExcel.xlsx')

