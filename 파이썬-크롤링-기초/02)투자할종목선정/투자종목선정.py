import openpyxl
import pandas as pd
import numpy as np
import getpass
from pykrx import stock

wb = openpyxl.Workbook()
ws = wb.create_sheet('현재가')

attrs = [
    '종목코드'
    , '종목명'
]

for idx, val in enumerate(attrs):
    ws[f'{chr(65+idx)}1'] = val

tickers = stock.get_market_ticker_list()
# print(tickers)
for idx, ticker in enumerate(tickers,2):
        종목 = stock.get_market_ticker_name(ticker)
        ws[f'A{idx}'] = ticker
        ws[f'B{idx}'] = 종목
        # print(종목)

wb.save('testExcel.xlsx')
